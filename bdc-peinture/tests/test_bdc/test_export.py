"""
Tests unitaires — export facturation (service Excel, vue, templates).
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from apps.bdc.exports import COLONNES, generer_export_excel
from apps.bdc.models import BonDeCommande, StatutChoices

# ─── Fixtures ────────────────────────────────────────────────────────────────


@pytest.fixture
def bdc_a_facturer(db, bailleur_gdh, utilisateur_cdt, sous_traitant):
    return BonDeCommande.objects.create(
        numero_bdc="EXP-001",
        bailleur=bailleur_gdh,
        adresse="10 Rue Export",
        ville="Avignon",
        occupation="OCCUPE",
        statut=StatutChoices.A_FACTURER,
        sous_traitant=sous_traitant,
        pourcentage_st=Decimal("65"),
        montant_ht=Decimal("1000"),
        montant_st=Decimal("650"),
        date_realisation=date(2026, 2, 15),
        cree_par=utilisateur_cdt,
    )


@pytest.fixture
def bdc_facture(db, bailleur_gdh, utilisateur_cdt, sous_traitant):
    return BonDeCommande.objects.create(
        numero_bdc="EXP-002",
        bailleur=bailleur_gdh,
        adresse="20 Rue Export",
        ville="Orange",
        occupation="VACANT",
        statut=StatutChoices.FACTURE,
        sous_traitant=sous_traitant,
        pourcentage_st=Decimal("60"),
        montant_ht=Decimal("2000"),
        montant_st=Decimal("1200"),
        date_realisation=date(2026, 1, 10),
        cree_par=utilisateur_cdt,
    )


@pytest.fixture
def bdc_en_cours(db, bailleur_gdh, utilisateur_cdt, sous_traitant):
    """BDC EN_COURS — ne doit pas apparaître dans l'export par défaut."""
    return BonDeCommande.objects.create(
        numero_bdc="EXP-003",
        bailleur=bailleur_gdh,
        adresse="30 Rue Export",
        ville="Avignon",
        occupation="OCCUPE",
        statut=StatutChoices.EN_COURS,
        sous_traitant=sous_traitant,
        pourcentage_st=Decimal("50"),
        cree_par=utilisateur_cdt,
    )


def _load_wb(response):
    """Helper : charge un workbook depuis une HttpResponse."""
    return load_workbook(BytesIO(response.content))


# ─── 7.1 Tests service generer_export_excel ──────────────────────────────────


class TestGenererExportExcel:
    def test_colonnes_correctes(self, bdc_a_facturer):
        qs = BonDeCommande.objects.filter(pk=bdc_a_facturer.pk)
        response = generer_export_excel(qs)
        wb = _load_wb(response)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == COLONNES

    def test_donnees_presentes(self, bdc_a_facturer):
        qs = BonDeCommande.objects.filter(pk=bdc_a_facturer.pk)
        response = generer_export_excel(qs)
        wb = _load_wb(response)
        ws = wb.active
        row = [cell.value for cell in ws[2]]
        assert row[0] == "EXP-001"
        assert row[2] == "10 Rue Export"
        assert row[3] == "Avignon"
        assert row[5] == 65.0  # % ST
        assert row[6] == 1000.0  # Montant HT
        assert row[7] == 650.0  # Montant ST
        assert row[8] == "15/02/2026"

    def test_fichier_xlsx_valide(self, bdc_a_facturer):
        qs = BonDeCommande.objects.filter(pk=bdc_a_facturer.pk)
        response = generer_export_excel(qs)
        assert response["Content-Type"] == ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert "export_facturation_" in response["Content-Disposition"]
        assert ".xlsx" in response["Content-Disposition"]

    def test_nom_fichier_date_du_jour(self, bdc_a_facturer):
        qs = BonDeCommande.objects.filter(pk=bdc_a_facturer.pk)
        response = generer_export_excel(qs)
        expected = f"export_facturation_{date.today().isoformat()}.xlsx"
        assert expected in response["Content-Disposition"]

    def test_export_vide(self, db):
        qs = BonDeCommande.objects.none()
        response = generer_export_excel(qs)
        wb = _load_wb(response)
        ws = wb.active
        assert ws.max_row == 1  # Seulement l'en-tête

    def test_headers_en_gras(self, bdc_a_facturer):
        qs = BonDeCommande.objects.filter(pk=bdc_a_facturer.pk)
        response = generer_export_excel(qs)
        wb = _load_wb(response)
        ws = wb.active
        for cell in ws[1]:
            assert cell.font.bold is True

    def test_plusieurs_lignes(self, bdc_a_facturer, bdc_facture):
        qs = BonDeCommande.objects.filter(pk__in=[bdc_a_facturer.pk, bdc_facture.pk])
        response = generer_export_excel(qs)
        wb = _load_wb(response)
        ws = wb.active
        assert ws.max_row == 3  # 1 en-tête + 2 lignes


# ─── 7.2 Tests vue export_facturation ────────────────────────────────────────


URL_EXPORT = "bdc:export_facturation"


class TestVueExportAcces:
    def test_cdt_acces_ok(self, client, utilisateur_cdt, bdc_a_facturer):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse(URL_EXPORT))
        assert response.status_code == 200

    def test_secretaire_can_access(self, client, utilisateur_secretaire):
        client.force_login(utilisateur_secretaire)
        response = client.get(reverse(URL_EXPORT))
        assert response.status_code == 200

    def test_anonyme_redirige(self, client):
        response = client.get(reverse(URL_EXPORT))
        assert response.status_code == 302


class TestVueExportFiltres:
    def test_filtre_statut_a_facturer(self, client, utilisateur_cdt, bdc_a_facturer, bdc_facture):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse(URL_EXPORT), {"statut": "A_FACTURER"})
        assert response.context["count"] == 1

    def test_filtre_statut_facture(self, client, utilisateur_cdt, bdc_a_facturer, bdc_facture):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse(URL_EXPORT), {"statut": "FACTURE"})
        assert response.context["count"] == 1

    def test_filtre_tous_statuts(self, client, utilisateur_cdt, bdc_a_facturer, bdc_facture):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse(URL_EXPORT))
        assert response.context["count"] == 2

    def test_filtre_sous_traitant(self, client, utilisateur_cdt, bdc_a_facturer, sous_traitant):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse(URL_EXPORT), {"sous_traitant": sous_traitant.pk})
        assert response.context["count"] == 1

    def test_filtre_date_du(self, client, utilisateur_cdt, bdc_a_facturer, bdc_facture):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse(URL_EXPORT), {"date_du": "2026-02-01"})
        assert response.context["count"] == 1  # Seulement EXP-001

    def test_filtre_date_au(self, client, utilisateur_cdt, bdc_a_facturer, bdc_facture):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse(URL_EXPORT), {"date_au": "2026-01-31"})
        assert response.context["count"] == 1  # Seulement EXP-002

    def test_en_cours_exclus_par_defaut(self, client, utilisateur_cdt, bdc_en_cours):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse(URL_EXPORT))
        assert response.context["count"] == 0


class TestVueExportTelechargement:
    def test_post_telecharge_xlsx(self, client, utilisateur_cdt, bdc_a_facturer):
        client.force_login(utilisateur_cdt)
        response = client.post(reverse(URL_EXPORT))
        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]

    def test_post_avec_filtres_get(self, client, utilisateur_cdt, bdc_a_facturer, bdc_facture):
        client.force_login(utilisateur_cdt)
        url = reverse(URL_EXPORT) + "?statut=A_FACTURER"
        response = client.post(url)
        wb = _load_wb(response)
        ws = wb.active
        assert ws.max_row == 2  # En-tête + 1 ligne A_FACTURER


# ─── 7.3 Tests template ─────────────────────────────────────────────────────


class TestTemplateExport:
    def test_apercu_compte_affiche(self, client, utilisateur_cdt, bdc_a_facturer, bdc_facture):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse(URL_EXPORT))
        content = response.content.decode()
        assert "2" in content  # 2 BDC
        assert "Télécharger" in content

    def test_bouton_desactive_si_zero(self, client, utilisateur_cdt):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse(URL_EXPORT))
        content = response.content.decode()
        assert "disabled" in content
        assert "Aucun BDC" in content


class TestTemplateRecoupementExporter:
    def test_bouton_exporter_visible(self, client, utilisateur_cdt, bdc_a_facturer):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse("bdc:recoupement_liste"))
        content = response.content.decode()
        assert "Exporter" in content
        assert reverse("bdc:export_facturation") in content


class TestTemplateDashboardExport:
    def test_lien_export_visible_cdt(self, client, utilisateur_cdt, bdc_a_facturer):
        client.force_login(utilisateur_cdt)
        response = client.get(reverse("bdc:index"))
        content = response.content.decode()
        assert "Export facturation" in content
        assert reverse("bdc:export_facturation") in content

    def test_lien_export_visible_secretaire(self, client, utilisateur_secretaire, bdc_a_facturer):
        client.force_login(utilisateur_secretaire)
        response = client.get(reverse("bdc:index"))
        content = response.content.decode()
        assert "Export facturation" in content
