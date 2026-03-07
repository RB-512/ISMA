"""Tests unitaires — relevés de facturation sous-traitant."""

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from apps.bdc.models import BonDeCommande, ReleveFacturation, ReleveStatutChoices, StatutChoices
from apps.bdc.releves import (
    ReleveError,
    creer_releve,
    retirer_bdc_du_releve,
    valider_releve,
)
from apps.bdc.releves_export import generer_releve_excel, generer_releve_pdf

# ─── Fixtures ────────────────────────────────────────────────────────────────


@pytest.fixture
def bdc_en_cours(db, bailleur_gdh, utilisateur_cdt, sous_traitant):
    return BonDeCommande.objects.create(
        numero_bdc="REL-001",
        bailleur=bailleur_gdh,
        adresse="10 Rue Releve",
        ville="Avignon",
        occupation="OCCUPE",
        statut=StatutChoices.A_FACTURER,
        sous_traitant=sous_traitant,
        pourcentage_st=Decimal("65"),
        montant_ht=Decimal("1000"),
        montant_st=Decimal("650"),
        date_realisation=date(2026, 2, 15),
        cree_par=utilisateur_cdt,
    )


@pytest.fixture
def bdc_facture(db, bailleur_gdh, utilisateur_cdt, sous_traitant):
    return BonDeCommande.objects.create(
        numero_bdc="REL-002",
        bailleur=bailleur_gdh,
        adresse="20 Rue Releve",
        ville="Orange",
        occupation="VACANT",
        statut=StatutChoices.FACTURE,
        sous_traitant=sous_traitant,
        pourcentage_st=Decimal("60"),
        montant_ht=Decimal("2000"),
        montant_st=Decimal("1200"),
        date_realisation=date(2026, 1, 10),
        cree_par=utilisateur_cdt,
    )


# ─── 1. Tests modele ────────────────────────────────────────────────────────


class TestReleveFacturationModel:
    def test_creation_releve(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = ReleveFacturation.objects.create(
            numero=1,
            sous_traitant=sous_traitant,
            cree_par=utilisateur_cdt,
        )
        releve.bdc.add(bdc_en_cours)
        assert releve.statut == ReleveStatutChoices.BROUILLON
        assert releve.numero == 1
        assert releve.bdc.count() == 1

    def test_str_representation(self, db, sous_traitant, utilisateur_cdt):
        releve = ReleveFacturation.objects.create(
            numero=3,
            sous_traitant=sous_traitant,
            cree_par=utilisateur_cdt,
        )
        assert "Dupont Peinture" in str(releve)
        assert "n°3" in str(releve)

    def test_montant_total(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours, bdc_facture):
        releve = ReleveFacturation.objects.create(
            numero=1,
            sous_traitant=sous_traitant,
            cree_par=utilisateur_cdt,
        )
        releve.bdc.add(bdc_en_cours, bdc_facture)
        assert releve.montant_total == Decimal("1850")  # 650 + 1200

    def test_montant_total_vide(self, db, sous_traitant, utilisateur_cdt):
        releve = ReleveFacturation.objects.create(
            numero=1,
            sous_traitant=sous_traitant,
            cree_par=utilisateur_cdt,
        )
        assert releve.montant_total == Decimal("0")

    def test_nb_bdc(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours, bdc_facture):
        releve = ReleveFacturation.objects.create(
            numero=1,
            sous_traitant=sous_traitant,
            cree_par=utilisateur_cdt,
        )
        releve.bdc.add(bdc_en_cours, bdc_facture)
        assert releve.nb_bdc == 2

    def test_periode(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours, bdc_facture):
        releve = ReleveFacturation.objects.create(
            numero=1,
            sous_traitant=sous_traitant,
            cree_par=utilisateur_cdt,
        )
        releve.bdc.add(bdc_en_cours, bdc_facture)
        debut, fin = releve.periode
        assert debut == date(2026, 1, 10)
        assert fin == date(2026, 2, 15)

    def test_periode_vide(self, db, sous_traitant, utilisateur_cdt):
        releve = ReleveFacturation.objects.create(
            numero=1,
            sous_traitant=sous_traitant,
            cree_par=utilisateur_cdt,
        )
        debut, fin = releve.periode
        assert debut is None
        assert fin is None

    def test_statut_defaut_brouillon(self, db, sous_traitant, utilisateur_cdt):
        releve = ReleveFacturation.objects.create(
            numero=1,
            sous_traitant=sous_traitant,
            cree_par=utilisateur_cdt,
        )
        assert releve.statut == ReleveStatutChoices.BROUILLON
        assert releve.date_validation is None


# ─── 2. Tests service creer_releve ──────────────────────────────────────────


class TestCreerReleve:
    def test_creer_releve_recupere_bdc_eligibles(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours, bdc_facture):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        assert releve.statut == ReleveStatutChoices.BROUILLON
        assert releve.bdc.count() == 2
        assert releve.numero == 1

    def test_numero_auto_incremente(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours, bailleur_gdh):
        r1 = creer_releve(sous_traitant, utilisateur_cdt)
        valider_releve(r1, utilisateur_cdt)
        # Creer un nouveau BDC eligible
        BonDeCommande.objects.create(
            numero_bdc="REL-003",
            bailleur=bailleur_gdh,
            adresse="30 Rue Test",
            ville="Avignon",
            statut=StatutChoices.A_FACTURER,
            sous_traitant=sous_traitant,
            montant_st=Decimal("100"),
            date_realisation=date(2026, 3, 1),
            cree_par=utilisateur_cdt,
        )
        r2 = creer_releve(sous_traitant, utilisateur_cdt)
        assert r2.numero == 2

    def test_exclut_bdc_deja_dans_releve_valide(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours, bdc_facture):
        r1 = creer_releve(sous_traitant, utilisateur_cdt)
        valider_releve(r1, utilisateur_cdt)
        with pytest.raises(ReleveError, match="Aucun BDC"):
            creer_releve(sous_traitant, utilisateur_cdt)

    def test_erreur_si_aucun_bdc_eligible(self, db, sous_traitant, utilisateur_cdt):
        with pytest.raises(ReleveError, match="Aucun BDC"):
            creer_releve(sous_traitant, utilisateur_cdt)

    def test_exclut_bdc_en_cours_statut(self, db, bailleur_gdh, sous_traitant, utilisateur_cdt):
        """Les BDC EN_COURS ne sont pas eligibles (seulement A_FACTURER + FACTURE)."""
        BonDeCommande.objects.create(
            numero_bdc="REL-EN-COURS",
            bailleur=bailleur_gdh,
            adresse="1 Rue Test",
            ville="Avignon",
            statut=StatutChoices.EN_COURS,
            sous_traitant=sous_traitant,
            cree_par=utilisateur_cdt,
        )
        with pytest.raises(ReleveError, match="Aucun BDC"):
            creer_releve(sous_traitant, utilisateur_cdt)


# ─── 3. Tests service retirer_bdc_du_releve ─────────────────────────────────


class TestRetirerBdc:
    def test_retirer_bdc_du_brouillon(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours, bdc_facture):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        assert releve.bdc.count() == 2
        retirer_bdc_du_releve(releve, bdc_en_cours)
        assert releve.bdc.count() == 1

    def test_retirer_bdc_releve_valide_interdit(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        valider_releve(releve, utilisateur_cdt)
        with pytest.raises(ReleveError, match="validé"):
            retirer_bdc_du_releve(releve, bdc_en_cours)


# ─── 4. Tests service valider_releve ────────────────────────────────────────


class TestValiderReleve:
    def test_valider_passe_en_valide(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        valider_releve(releve, utilisateur_cdt)
        releve.refresh_from_db()
        assert releve.statut == ReleveStatutChoices.VALIDE
        assert releve.date_validation is not None

    def test_valider_releve_deja_valide_interdit(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        valider_releve(releve, utilisateur_cdt)
        with pytest.raises(ReleveError, match="déjà validé"):
            valider_releve(releve, utilisateur_cdt)

    def test_valider_releve_vide_interdit(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        releve.bdc.clear()
        with pytest.raises(ReleveError, match="vide"):
            valider_releve(releve, utilisateur_cdt)

    def test_anti_doublon_apres_validation(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        """Un BDC dans un relevé validé ne peut pas être dans un nouveau relevé."""
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        valider_releve(releve, utilisateur_cdt)
        with pytest.raises(ReleveError, match="Aucun BDC"):
            creer_releve(sous_traitant, utilisateur_cdt)


# ─── 5. Tests vues ──────────────────────────────────────────────────────────


class TestVueCreerReleve:
    def test_creer_releve_post(self, client_cdt, sous_traitant, bdc_en_cours):
        url = reverse("bdc:releve_creer", kwargs={"st_pk": sous_traitant.pk})
        response = client_cdt.post(url)
        assert response.status_code == 302
        assert ReleveFacturation.objects.count() == 1

    def test_creer_releve_sans_bdc_eligible(self, client_cdt, sous_traitant):
        url = reverse("bdc:releve_creer", kwargs={"st_pk": sous_traitant.pk})
        response = client_cdt.post(url)
        assert response.status_code == 302
        assert ReleveFacturation.objects.count() == 0

    def test_anonyme_redirige(self, client, sous_traitant):
        url = reverse("bdc:releve_creer", kwargs={"st_pk": sous_traitant.pk})
        response = client.post(url)
        assert response.status_code == 302
        assert "/accounts/" in response.url


class TestVueReleveDetail:
    def test_detail_brouillon(self, client_cdt, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        url = reverse("bdc:releve_detail", kwargs={"pk": releve.pk})
        response = client_cdt.get(url)
        assert response.status_code == 200
        content = response.content.decode()
        assert "REL-001" in content
        assert "Brouillon" in content

    def test_detail_valide(self, client_cdt, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        valider_releve(releve, utilisateur_cdt)
        url = reverse("bdc:releve_detail", kwargs={"pk": releve.pk})
        response = client_cdt.get(url)
        assert response.status_code == 200
        content = response.content.decode()
        assert "Valid" in content


class TestVueValiderReleve:
    def test_valider_post(self, client_cdt, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        url = reverse("bdc:releve_valider", kwargs={"pk": releve.pk})
        response = client_cdt.post(url)
        assert response.status_code == 302
        releve.refresh_from_db()
        assert releve.statut == ReleveStatutChoices.VALIDE


class TestVueRetirerBdc:
    def test_retirer_bdc_post(self, client_cdt, sous_traitant, utilisateur_cdt, bdc_en_cours, bdc_facture):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        url = reverse("bdc:releve_retirer_bdc", kwargs={"pk": releve.pk, "bdc_pk": bdc_en_cours.pk})
        response = client_cdt.post(url)
        assert response.status_code == 302
        assert releve.bdc.count() == 1


class TestVueHistoriqueReleves:
    def test_historique_st(self, client_cdt, sous_traitant, utilisateur_cdt, bdc_en_cours):
        creer_releve(sous_traitant, utilisateur_cdt)
        url = reverse("bdc:releve_historique", kwargs={"st_pk": sous_traitant.pk})
        response = client_cdt.get(url)
        assert response.status_code == 200
        content = response.content.decode()
        assert "1" in content


# ─── 6. Tests export PDF ────────────────────────────────────────────────────


class TestRelevePdf:
    def test_pdf_genere(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        response = generer_releve_pdf(releve)
        assert response.status_code == 200
        assert response["Content-Type"] == "application/pdf"
        assert "releve_" in response["Content-Disposition"]

    def test_pdf_contient_donnees(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        response = generer_releve_pdf(releve)
        assert len(response.content) > 100


# ─── 7. Tests export Excel ──────────────────────────────────────────────────


class TestReleveExcel:
    def test_excel_genere(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        response = generer_releve_excel(releve)
        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]

    def test_excel_contient_bdc(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours, bdc_facture):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        response = generer_releve_excel(releve)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.max_row == 3  # header + 2 BDC

    def test_excel_colonnes(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        response = generer_releve_excel(releve)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "N\u00b0 BDC" in headers
        assert "Montant ST (\u20ac)" in headers

    def test_excel_headers_en_gras(self, db, sous_traitant, utilisateur_cdt, bdc_en_cours):
        releve = creer_releve(sous_traitant, utilisateur_cdt)
        response = generer_releve_excel(releve)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        for cell in ws[1]:
            assert cell.font.bold is True


# ─── 8. Tests integration recoupement ────────────────────────────────────────


class TestRecoupementIntegration:
    def test_bouton_nouveau_releve_visible(self, client_cdt, sous_traitant, bdc_en_cours):
        response = client_cdt.get(reverse("bdc:recoupement_liste"))
        content = response.content.decode()
        assert "Nouveau relev" in content or "releve_creer" in content

    def test_lien_historique_visible(self, client_cdt, sous_traitant, bdc_en_cours):
        response = client_cdt.get(reverse("bdc:recoupement_liste"))
        content = response.content.decode()
        assert "releve_historique" in content or "Relev" in content
